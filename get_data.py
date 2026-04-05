import openpyxl

# Mở file Excel
wb = openpyxl.load_workbook("8D.xlsx")
sheet = wb.active

import openpyxl

wb = openpyxl.load_workbook("8D.xlsx")
sheet = wb.active

with open("output.txt", "w", encoding="utf-8") as f:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        ma_ts = row[0]   # cột A (8D-01)
        ten_ts = row[1]  # cột B (TÊN IN HOA)

        if ma_ts and ten_ts:
            # 👉 Bỏ "8D-"
            stt = ma_ts.split("-")[1]  # lấy phần sau dấu -

            # 👉 Bỏ số 0 đầu (01 → 1)
            stt = str(int(stt))

            # 👉 Viết hoa chữ cái đầu mỗi từ
            ten_ts = ten_ts.lower().title()

            # 👉 Ghi ra file
            f.write(f"{stt}. {ten_ts}\n")